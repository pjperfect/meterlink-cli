import openpyxl, os
from time import sleep
from storage import load_purchases, save_purchases, get_next_id


def parse_mpesa_message(full_message_text):
    """
    Example:
    TLD910G82S Confirmed. Ksh1,000.00 sent to KPLC PREPAID for account 04215776788 on 13/12/25 at 11:18 PM ...
    We extract:
    - mpesa_code (first word)
    - meter_number (after 'account ')
    - amount (after 'Ksh')
    - date_purchase (after ' on ' ... we keep it as text)
    """
    text = full_message_text.strip()

    # mpesa code
    parts = text.split() # parts = ["TLD910G82S", "Confirmed.", "Ksh1,000.00", "sent", "to", ...]
    mpesa_code = None
    if len(parts) > 0: # We check if len(parts) > 0 so we don’t crash if the message is empty.
        mpesa_code = parts[0]

    # meter number
    meter_number = None
    key = "account "
    account_position = text.find(key)
    if account_position != -1: # everything after account
        after_position = text[account_position + len(key):]
        meter_number = after_position.split(" ")[0].strip()

    #amount
    amount = None
    ksh_key = "Ksh"
    ksh_position = text.find(ksh_key)
    if ksh_position != -1:
        after_ksh = text[ksh_position + len(ksh_key):]
        amount_text = after_ksh.split(" ")[0].strip()
        amount_text = amount_text.replace(",", "")
        try:
            amount = float(amount_text)
        except:
            amount = None

    # date purchase
    date_purchase = None
    on_key = " on "
    on_position = text.find(on_key)
    if on_position != -1:
        after_on = text[on_position + len(on_key):]
        stop_key = " New "
        stop_position = after_on.find(stop_key)

        if stop_position == -1:
            date_part = after_on.strip()
        else:
            date_part = after_on[:stop_position].strip()

        date_part = date_part.replace(" at ", " ")
        date_purchase = date_part

    return {
        "meter_number": meter_number,
        "amount": amount,
        "date_purchase": date_purchase,
        "mpesa_code": mpesa_code
    }


def parse_kplc_message(full_message_text):
    """
    Example KPLC message:
    Mtr:04215776788
    Token:0139-7235-7028-2532-0608
    Date:20251213 23:19
    Units:39.8
    Amt:1000.00
    TknAmt:656.21
    OtherCharges:343.79
    What we want:
    - meter_number: from "Mtr:"
    - token:        from "Token:"
    - date_purchase:from "Date:"
    - amount:       from "Amt:"
    """
    text = full_message_text.strip()

    #break message into lines
    lines = text.splitlines()
    message_data = {}

    for line in lines:
        line = line.strip()
        if ":" in line:
            left_side, right_side = line.split(":", 1)
            left_side = left_side.strip()
            right_side = right_side.strip()
            message_data[left_side] = right_side

    meter_number = message_data.get("Mtr")
    token = message_data.get("Token")
    date_purchase = message_data.get("Date")

    amount = None
    amount_text = message_data.get("Amt")
    if amount_text is not None:
        try:
            amount = float(amount_text)
        except:
            amount = None

    return {
        "meter_number": meter_number,
        "token": token,
        "date_purchase": date_purchase,
        "amount": amount
    }


def import_excel_for_meter(selected_meter):

    if selected_meter is None:
        print("Select a meter first.")
        sleep(1)
        return 0
    
    print("\n---- Import Excel ----")
    print(f"Selected meter: {selected_meter['alias']}")

    default_file = "meterlink_template.xlsx"

    # LOOP: keep asking for file name until it opens (or user cancels)
    while True:

        print("\nType 0 to cancel.")
        user_input = input("Excel file name (press Enter for template): ").strip()

        if user_input == "0":
            print("Cancelled.")
            sleep(1)
            return 0

        if user_input == "":
            user_input = default_file

        if user_input != default_file:
            if not user_input.lower().endswith(".xlsx"):
                user_input = user_input + ".xlsx"

        file_path = os.path.join("templates", user_input)

        if not os.path.exists(file_path):
            print("File not found.")
            print("Put the Excel file inside 'templates/'.")
            sleep(1)
            continue

        try:
            workbook = openpyxl.load_workbook(file_path)
            excel_sheet = workbook.active
            break
        except:
            print("Could not open Excel file.")
            print("Make sure the file is closed, then try again.")
            sleep(1)
            continue


    purchases_list = load_purchases()
    imported_count = 0
    skipped_count = 0

    # replace old purchases for this meter (must be y or n)
    while True:
        choice = input("Replace existing purchases for this meter? (y/n): ").strip().lower()

        if choice == "y":
            kept_purchases = []
            for purchase in purchases_list:
                if purchase["meter_id"] != selected_meter["id"]:
                    kept_purchases.append(purchase)

            purchases_list = kept_purchases
            print("Old purchases removed for this meter.")
            sleep(1)
            break

        if choice == "n":
            print("Keeping existing purchases. Import will add on top.")
            sleep(1)
            break

        print("Invalid choice. Type y or n.")
        sleep(1)

    row = 3
    while True:

        mpesa_cell_value = excel_sheet.cell(row = row, column = 1).value
        kplc_cell_value = excel_sheet.cell(row = row, column = 2).value

        if mpesa_cell_value is None and kplc_cell_value is None:
            break

        mpesa_text = ""
        if mpesa_cell_value is not None:
            mpesa_text = str(mpesa_cell_value).strip()

        kplc_text = ""
        if kplc_cell_value is not None:
            kplc_text = str(kplc_cell_value).strip()

        # if both are blank, skip
        if mpesa_text == "" and kplc_text == "":
            skipped_count = skipped_count + 1
            row = row + 1
            continue

        mpesa_data = None
        if mpesa_text != "":
            mpesa_data = parse_mpesa_message(mpesa_text)

        kplc_data = None
        if kplc_text != "":
            kplc_data = parse_kplc_message(kplc_text)


        # meter check (skip if message meter doesn't match selected meter)
        selected_meter_number = str(selected_meter["meter_number"])

        if mpesa_data is not None:
            if mpesa_data["meter_number"] is not None:
                if str(mpesa_data["meter_number"]) != selected_meter_number:
                    skipped_count = skipped_count + 1
                    row = row + 1
                    continue

        if kplc_data is not None:
            if kplc_data["meter_number"] is not None:
                if str(kplc_data["meter_number"]) != selected_meter_number:
                    skipped_count = skipped_count + 1
                    row = row + 1
                    continue


        # choose amount and date (prefer KPLC, fallback to M-Pesa)
        amount = None
        date_purchase = None

        if kplc_data is not None:
            amount = kplc_data["amount"]
            date_purchase = kplc_data["date_purchase"]

        if amount is None and mpesa_data is not None:
            amount = mpesa_data["amount"]

        if date_purchase is None and mpesa_data is not None:
            date_purchase = mpesa_data["date_purchase"]


        # choose code (prefer mpesa code, else token)
        code = "N/A"
        if mpesa_data is not None and mpesa_data["mpesa_code"] is not None:
            code = mpesa_data["mpesa_code"]
        elif kplc_data is not None and kplc_data["token"] is not None:
            code = kplc_data["token"]


        # validate amount and date
        if amount is None or date_purchase is None:
            skipped_count = skipped_count + 1
            row = row + 1
            continue

        try:
            amount_int = int(float(amount))
        except:
            skipped_count = skipped_count + 1
            row = row + 1
            continue

        if amount_int <= 0:
            skipped_count = skipped_count + 1
            row = row + 1
            continue


        # combine messages for saving
        combined_message = ""
        if mpesa_text != "":
            combined_message = mpesa_text

        if kplc_text != "":
            if combined_message != "":
                combined_message = combined_message + "\n\n"
            combined_message = combined_message + kplc_text


        new_purchase = {
            "id": get_next_id(purchases_list),
            "meter_id": selected_meter["id"],
            "amount": amount_int,
            "date_purchase": str(date_purchase),
            "mpesa_code": str(code),
            "message": combined_message
        }

        purchases_list.append(new_purchase)
        imported_count = imported_count + 1
        row = row + 1


    save_purchases(purchases_list)
    print(f"Successfully Imported {imported_count} purchases.")
    print(f"Skipped {skipped_count} rows.")
    sleep(1)
    return imported_count